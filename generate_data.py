# -*- coding: utf-8 -*-
import openpyxl
import random
import datetime

random.seed(2026)

# ========== Load source ==========
wb = openpyxl.load_workbook('.dumate/inbox/国投看板低保真0322.xlsx')
ws = wb['点击项目数字的弹窗显示列']

# ========== Helper: resolve merged cell value ==========
def get_merged_value(ws, row, col):
    """Get value from a cell, resolving merged cell ranges."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    # Check if this cell is part of a merged range
    for mc in ws.merged_cells.ranges:
        if cell.coordinate in mc:
            # Return the top-left cell's value
            return ws.cell(row=mc.min_row, column=mc.min_col).value
    return None

# ========== Identify data rows vs subtotal rows ==========
data_rows = []
for r in range(3, 53):
    b_val = get_merged_value(ws, r, 2)
    if b_val == '小计' or b_val == '合计':
        continue
    data_rows.append(r)

num_data = len(data_rows)

# ========== Supplier pool (10 suppliers) ==========
suppliers = [
    '中航信通科技有限公司',
    '华宇信息技术股份有限公司',
    '东软集团股份有限公司',
    '浪潮电子信息产业股份有限公司',
    '神州数码信息服务股份有限公司',
    '太极计算机股份有限公司',
    '东华软件股份公司',
    '用友网络科技股份有限公司',
    '金蝶软件（中国）有限公司',
    '中科软科技股份有限公司',
]

max_per_supplier = int(num_data * 0.3)  # 12
supplier_assignments = []
supplier_counts = {s: 0 for s in suppliers}
for i in range(num_data):
    available = [s for s in suppliers if supplier_counts[s] < max_per_supplier]
    min_count = min(supplier_counts[s] for s in available)
    candidates = [s for s in available if supplier_counts[s] == min_count]
    chosen = random.choice(candidates)
    supplier_assignments.append(chosen)
    supplier_counts[chosen] += 1

# ========== Project name generation ==========
project_name_parts = {
    'A0201': ['财务核算', '业财一体化', '会计核算', '总账管理'],
    'A0202': ['财务合并', '报表合并', '集团合并报表', '合并抵销'],
    'A0203': ['全面预算', '预算管理', '预算编制', '预算执行监控'],
    'A0204': ['司库管理', '资金管理', '银企直联', '现金流管理'],
    'A0205': ['境外资金', '跨境资金', '外汇管理', '境外资金池'],
    'A0206': ['财务共享', '共享中心', '费用报账', '影像档案'],
    'A0207': ['税务管理', '发票管理', '税务申报', '税务风控'],
    'A0209': ['财务数据', '数据综合', '财务数据中台', '数据治理'],
    'A0301': ['人资数字化', '人力资源', '人才管理', '人事管理'],
    'A0302': ['网上大学', '在线学习', '培训管理', '学习平台'],
    'A0303': ['党建云', '智慧党建', '党员管理', '党建平台'],
    'A0304': ['数字媒体', '新媒体管理', '内容管理', '融媒体'],
    'INFRA': ['基础设施', '系统集成', '网络建设', '服务器部署', '存储系统', '云平台', '数据中心', '网络运维'],
}

project_type_words = {
    '新建': ['系统建设项目', '平台建设工程', '系统搭建项目', '信息化建设项目'],
    '优化': ['系统升级改造', '功能优化项目', '系统增强工程', '迭代优化项目'],
    '运维（含安全服务）': ['运维服务项目', '安全运维项目', '系统运维保障', '安全服务项目'],
    '咨询': ['咨询服务项目', '规划咨询项目', '可行性研究项目', '评估咨询项目'],
}

phase_suffix = ['一期', '二期', '三期', '四期', '', '', '']
used_names = set()

def get_system_code(b_val):
    """Extract system code from system name string."""
    b_str = str(b_val) if b_val else ''
    if b_str.startswith('A02') or b_str.startswith('A03'):
        return b_str[:5].strip()
    return 'INFRA'

def gen_project_name(system_code, project_type):
    parts = project_name_parts.get(system_code, project_name_parts['INFRA'])
    type_words = project_type_words.get(project_type, ['项目'])
    for _ in range(20):
        name = random.choice(parts) + random.choice(type_words)
        suffix = random.choice(phase_suffix)
        if suffix:
            name += suffix
        if name not in used_names:
            used_names.add(name)
            return name
    name = random.choice(parts) + random.choice(type_words) + str(random.randint(1, 999))
    used_names.add(name)
    return name

# ========== Phase assignment ==========
phase_by_type = {
    '新建': ['立项审批', '需求分析', '系统设计', '开发实施', '测试验收', '上线试运行'],
    '优化': ['需求分析', '系统设计', '开发实施', '测试验收', '上线试运行'],
    '运维（含安全服务）': ['运维实施', '运维验收', '运维完成'],
    '咨询': ['需求调研', '方案设计', '方案评审', '交付验收'],
}

phase_payment = {
    '立项审批':   (0.0, 0.05, 0),
    '需求分析':   (0.10, 0.20, 1),
    '需求调研':   (0.10, 0.20, 1),
    '系统设计':   (0.20, 0.30, 1),
    '方案设计':   (0.20, 0.30, 1),
    '开发实施':   (0.30, 0.45, 2),
    '方案评审':   (0.30, 0.40, 2),
    '测试验收':   (0.40, 0.55, 2),
    '交付验收':   (0.40, 0.55, 2),
    '上线试运行': (0.45, 0.58, 3),
    '运维实施':   (0.30, 0.45, 1),
    '运维验收':   (0.45, 0.55, 2),
    '运维完成':   (0.50, 0.58, 3),
    '项目评估':   (0.10, 0.30, 1),
    '新建':       (0.0, 0.10, 0),
}

def gen_contract_amount():
    base = random.choice([50, 80, 100, 150, 200, 300, 500, 800])
    amount = base * 10000 + random.randint(0, 9999)
    amount = round(amount / 1000) * 1000
    return amount

def add_months(dt, months):
    month = dt.month + months
    year = dt.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    day = min(dt.day, 28)
    return datetime.datetime(year, month, day)

def gen_dates(phase, project_type):
    sign_year = random.choice([2023, 2023, 2024, 2024])
    sign_month = random.randint(1, 12)
    sign_day = random.randint(1, 28)
    sign_date = datetime.datetime(sign_year, sign_month, sign_day)
    launch_months = random.randint(3, 12)
    launch_date = add_months(sign_date, launch_months)
    accept_months = random.randint(1, 6)
    accept_date = add_months(launch_date, accept_months)
    deadline = None
    if project_type == '运维（含安全服务）':
        deadline = add_months(sign_date, 12)
    early_phases = ['立项审批', '需求分析', '需求调研', '系统设计', '方案设计', '新建', '项目评估']
    if phase in early_phases:
        today = datetime.datetime(2025, 6, 30)
        if launch_date < today:
            launch_date = add_months(today, random.randint(1, 6))
            accept_date = add_months(launch_date, random.randint(1, 6))
    return sign_date, launch_date, accept_date, deadline

# ========== Process each data row ==========
for idx, r in enumerate(data_rows):
    b_val = get_merged_value(ws, r, 2)  # Resolve merged system name
    e_val = ws.cell(row=r, column=5).value
    project_type = e_val if e_val else '新建'
    system_code = get_system_code(b_val)
    
    # 1. Project name
    proj_name = gen_project_name(system_code, project_type)
    ws.cell(row=r, column=3, value=proj_name)
    
    # 2. Project phase
    if project_type in phase_by_type:
        new_phase = random.choice(phase_by_type[project_type])
    else:
        new_phase = random.choice(['需求分析', '开发实施', '测试验收'])
    ws.cell(row=r, column=6, value=new_phase)
    
    # 3. Contract stages (3~6)
    total_stages = random.choice([3, 4, 4, 4, 5, 6])
    ws.cell(row=r, column=7, value=total_stages)
    
    # 4. Contract amount
    contract_amount = gen_contract_amount()
    ws.cell(row=r, column=8, value=contract_amount)
    ws.cell(row=r, column=8).number_format = '#,##0'
    
    # 5. Dates
    sign_date, launch_date, accept_date, deadline = gen_dates(new_phase, project_type)
    ws.cell(row=r, column=9, value=sign_date)
    ws.cell(row=r, column=9).number_format = 'yyyy/m/d'
    ws.cell(row=r, column=10, value=launch_date)
    ws.cell(row=r, column=10).number_format = 'yyyy/m/d'
    ws.cell(row=r, column=11, value=accept_date)
    ws.cell(row=r, column=11).number_format = 'yyyy/m/d'
    if deadline:
        ws.cell(row=r, column=12, value=deadline)
        ws.cell(row=r, column=12).number_format = 'yyyy/m/d'
    
    # 6. Supplier
    ws.cell(row=r, column=13, value=supplier_assignments[idx])
    
    # 7. Paid/Unpaid
    if new_phase in phase_payment:
        min_ratio, max_ratio, typical_paid = phase_payment[new_phase]
    else:
        min_ratio, max_ratio, typical_paid = (0.1, 0.3, 1)
    
    paid_ratio = random.uniform(min_ratio, max_ratio)
    paid_ratio = min(paid_ratio, 0.58)
    
    paid_stages = min(typical_paid, total_stages - 1) if typical_paid > 0 else 0
    if paid_stages > 0 and random.random() < 0.3:
        paid_stages = max(0, paid_stages - 1)
    if paid_stages < total_stages - 1 and random.random() < 0.2:
        paid_stages = min(total_stages - 1, paid_stages + 1)
    
    unpaid_stages = total_stages - paid_stages
    
    paid_amount = round(contract_amount * paid_ratio / 1000) * 1000
    paid_amount = min(paid_amount, int(contract_amount * 0.6))
    
    ws.cell(row=r, column=14, value=paid_stages)
    ws.cell(row=r, column=15, value=unpaid_stages)
    ws.cell(row=r, column=16, value=paid_amount)
    ws.cell(row=r, column=16).number_format = '#,##0'
    ws.cell(row=r, column=17, value=f'=H{r}-P{r}')
    ws.cell(row=r, column=17).number_format = '#,##0'
    
    stage_amount = round(contract_amount / total_stages / 1000) * 1000
    ws.cell(row=r, column=18, value=stage_amount)
    ws.cell(row=r, column=18).number_format = '#,##0'

# ========== Fix subtotal SUM formulas ==========
groups = []
current_group = []
for r in range(3, 53):
    b_val = get_merged_value(ws, r, 2)
    if b_val == '小计':
        groups.append((r, current_group[:]))
        current_group = []
    elif b_val == '合计':
        pass
    else:
        current_group.append(r)

for sub_row, drows in groups:
    if not drows:
        continue
    first = drows[0]
    last = drows[-1]
    ws.cell(row=sub_row, column=8, value=f'=SUM(H{first}:H{last})')
    ws.cell(row=sub_row, column=16, value=f'=SUM(P{first}:P{last})')
    ws.cell(row=sub_row, column=17, value=f'=SUM(Q{first}:Q{last})')
    ws.cell(row=sub_row, column=18, value=f'=SUM(R{first}:R{last})')
    for col in [8, 16, 17, 18]:
        ws.cell(row=sub_row, column=col).number_format = '#,##0'

# Grand total (row 53)
all_first = data_rows[0]
all_last = data_rows[-1]
ws.cell(row=53, column=8, value=f'=SUM(H{all_first}:H{all_last})')
ws.cell(row=53, column=16, value=f'=SUM(P{all_first}:P{all_last})')
ws.cell(row=53, column=17, value=f'=SUM(Q{all_first}:Q{all_last})')
ws.cell(row=53, column=18, value=f'=SUM(R{all_first}:R{all_last})')
for col in [8, 16, 17, 18]:
    ws.cell(row=53, column=col).number_format = '#,##0'

# ========== Update 实施厂商分布 sheet ==========
ws2 = wb['实施厂商分布']

# Unmerge A9:B9
for mc in list(ws2.merged_cells.ranges):
    if 'A9' in str(mc) or 'B9' in str(mc):
        ws2.unmerge_cells(str(mc))

# Insert 4 rows before row 9 to get 10 supplier slots
ws2.insert_rows(9, amount=4)

# Write all 10 suppliers
for i, s in enumerate(suppliers):
    row_num = 3 + i
    ws2.cell(row=row_num, column=1, value=i + 1)
    ws2.cell(row=row_num, column=2, value=s)
    ws2.cell(row=row_num, column=3, value=f'=IF(B{row_num}="",0,SUMIFS(点击项目数字的弹窗显示列!H3:H52,点击项目数字的弹窗显示列!M3:M52,B{row_num},点击项目数字的弹窗显示列!B3:B52,"<>小计"))')
    ws2.cell(row=row_num, column=3).number_format = '#,##0'
    ws2.cell(row=row_num, column=4, value=f'=IFERROR(C{row_num}/C13,0)')
    ws2.cell(row=row_num, column=4).number_format = '0.00%'

# Update 合计 row (now row 13)
ws2.cell(row=13, column=1, value='合计')
ws2.cell(row=13, column=3, value='=SUM(C3:C12)')
ws2.cell(row=13, column=3).number_format = '#,##0'
ws2.cell(row=13, column=4, value='=IFERROR(C13/C13,0)')
ws2.cell(row=13, column=4).number_format = '0.00%'

# Re-merge A13:B13
ws2.merge_cells('A13:B13')

# ========== Save ==========
output_path = '国投测试数据.xlsx'
wb.save(output_path)
print(f"Saved to: {output_path}")

# ========== Verification ==========
print("\n========== Verification ==========")
wb2 = openpyxl.load_workbook(output_path, data_only=False)
ws_v = wb2['点击项目数字的弹窗显示列']

print("\nAll data rows:")
for r in data_rows:
    c_val = ws_v.cell(row=r, column=3).value
    e_val = ws_v.cell(row=r, column=5).value
    f_val = ws_v.cell(row=r, column=6).value
    g_val = ws_v.cell(row=r, column=7).value
    h_val = ws_v.cell(row=r, column=8).value
    m_val = ws_v.cell(row=r, column=13).value
    n_val = ws_v.cell(row=r, column=14).value
    o_val = ws_v.cell(row=r, column=15).value
    p_val = ws_v.cell(row=r, column=16).value
    r_val = ws_v.cell(row=r, column=18).value
    
    if h_val and isinstance(h_val, (int, float)):
        ratio = p_val / h_val if p_val and h_val else 0
        print(f"  R{r:2d}: {c_val} | {e_val} | {f_val} | 阶段={g_val} | 金额={h_val:>10,} | "
              f"{m_val[:8]:8s} | 付{n_val}/{g_val} | 已付={p_val:>10,} | 应付={r_val:>10,} | {ratio:.0%}")

# Paid ratio check
print("\n--- Paid ratio <= 60% check ---")
violations = 0
for r in data_rows:
    h_val = ws_v.cell(row=r, column=8).value
    p_val = ws_v.cell(row=r, column=16).value
    if h_val and p_val and isinstance(h_val, (int, float)) and isinstance(p_val, (int, float)):
        ratio = p_val / h_val
        if ratio > 0.60:
            print(f"  FAIL Row {r}: {p_val}/{h_val} = {ratio:.1%}")
            violations += 1
if violations == 0:
    print("  All 42 rows pass")

# Supplier distribution check
print("\n--- Supplier distribution (max 30%) ---")
supplier_in_data = {}
for r in data_rows:
    s = ws_v.cell(row=r, column=13).value
    if s:
        supplier_in_data[s] = supplier_in_data.get(s, 0) + 1
all_pass = True
for s, c in sorted(supplier_in_data.items(), key=lambda x: -x[1]):
    pct = c / len(data_rows) * 100
    status = "OK" if pct <= 30 else "FAIL"
    if pct > 30:
        all_pass = False
    print(f"  {s}: {c} projects ({pct:.1f}%) {status}")
if all_pass:
    print("  All suppliers within 30% limit")

# Project name uniqueness
print("\n--- Project name uniqueness ---")
names = [ws_v.cell(row=r, column=3).value for r in data_rows]
unique = len(set(names))
print(f"  {unique} unique names out of {len(names)} total")
if unique < len(names):
    dupes = set([n for n in names if names.count(n) > 1])
    print(f"  Duplicates: {dupes}")
else:
    print("  All names unique")

# Stage count check N+O=G
print("\n--- Stage count check (N+O = G) ---")
stage_ok = True
for r in data_rows:
    g = ws_v.cell(row=r, column=7).value
    n = ws_v.cell(row=r, column=14).value
    o = ws_v.cell(row=r, column=15).value
    if g and n is not None and o is not None:
        if n + o != g:
            print(f"  FAIL Row {r}: G={g}, N={n}, O={o}")
            stage_ok = False
if stage_ok:
    print("  All rows pass")

# Check same supplier for different projects
print("\n--- Supplier sharing check ---")
sup_to_rows = {}
for r in data_rows:
    s = ws_v.cell(row=r, column=13).value
    sup_to_rows.setdefault(s, []).append(r)
for s, rows in sup_to_rows.items():
    if len(rows) > 1:
        print(f"  {s}: {len(rows)} projects (rows {rows})")

