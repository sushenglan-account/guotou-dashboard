import json
import subprocess
import time
import sys

# Load schema reference
with open("schema_ref.txt") as f:
    schema_ref = f.read()

# Load questions
import openpyxl
wb = openpyxl.load_workbook(".dumate/inbox/问数调研问题收集.xlsx")
ws = wb.active

questions = []
for row in range(2, ws.max_row+1):
    a = ws.cell(row, 1).value or ""
    b = ws.cell(row, 2).value or ""
    c = ws.cell(row, 3).value or ""
    d = ws.cell(row, 4).value or ""
    e = ws.cell(row, 5).value or ""
    f_val = ws.cell(row, 6).value or ""
    g_val = ws.cell(row, 7).value or ""
    questions.append({
        "row": row,
        "A": str(a), "B": str(b), "C": str(c),
        "D": str(d), "E": str(e), "F": str(f_val), "G": str(g_val)
    })

print(f"Total questions: {len(questions)}")

# Split into batches of 8
batches = [questions[i:i+8] for i in range(0, len(questions), 8)]

API_KEY = "sk-w4ENEt8TfHvdBcYpD5411c8a173a41CfAb9404AbB5DeAdA6"
URL = "https://oneapi-comate.baidu-int.com/v1/messages"

all_results = {}

for batch_idx, batch in enumerate(batches):
    print(f"\n--- Batch {batch_idx+1}/{len(batches)} (rows {[q['row'] for q in batch]}) ---", flush=True)
    
    # Build questions text
    q_text = ""
    for q in batch:
        q_text += f"\n行号{q['row']}: 科室={q['B'][:30]}\n  问题: {q['C'][:200]}\n  现有D(数据属性资源名称): {q['D']}\n  现有E(数据表归属业务系统): {q['E']}\n  现有F(所属数据表): {q['F']}\n  现有G(所属字段): {q['G']}\n"
    
    prompt = f'''你是北京公交数据湖的资深数据分析师。请严格基于以下数据表schema信息，对每个用户问数问题进行分类判断。

{schema_ref}

=== 分类规则 ===

（1）已接入数据可回答：问题所需的数据字段在已接入的表中可以找到（数据湖指标表4张已接入 + 事件池11张已接入）。需要明确指出用哪张表、哪些字段。
  - H列填写："（1）已接入数据可回答"
  - F列填写：所属数据表名（已接入的表名）
  - G列填写：所属字段（多个字段用逗号分隔）

（2）未接入数据需要明确场景、统计口径和数据表：问题所需数据在数据湖1215张表或事件池28张表中存在对应的表，但这些表尚未接入。需要指出对应的未接入表名和可能用到的字段。
  - H列填写："（2）未接入数据需要明确场景、统计口径和数据表"
  - F列填写：建议接入的数据表名（未接入的表名）
  - G列填写：建议使用的字段（根据表名推断的关键字段）

（3）数据表schema不支持：问题所需的数据在数据湖和事件池的所有表中都不存在对应的数据字段，无法通过现有schema实现。
  - H列填写："（3）数据表schema不支持"
  - F列和G列留空

判断要点：
- 严格基于schema字段判断，不要臆造字段
- 已接入的表（可回答）优先判断，如果已接入的表能回答就归为（1）
- 如果已接入的表不能回答，但未接入的表中有对应数据，归为（2）
- 如果所有表都没有对应数据，归为（3）
- 主数据车辆维表(TD_MD_HX_CL_DA)包含：车型、燃料类型、分公司、车辆状态、牌照号等
- 主数据线路表(TD_MD_HX_XL_DA)包含：线路编号、线路名称、首末站、线路类型、票价等
- 主数据线路站点(TD_MD_HX_XL_ZD_DA)包含：站点名称、经纬度、站点序号、距下站距离等
- 主数据场站表(TI_MD_HX_CZ_DA)包含：场站名称、场站类型、经纬度、投产日期等
- 事件池主表(TA_SJC_MASTER_MI)包含：事件日期、车速、经纬度、最近的站位/城区/环路位置等
- 事件池-车辆运营路单(TA_SJC_SUB_BS_LD_DAY_MI)包含：路单、发车/终到、车速、经纬度等
- 事件池-充电交易表(TA_SJC_SUB_ZHJRQKL_CDJY_DETAIL)包含：充电交易、能源站等
- 事件池-主数据能源站信息表包含：能源站名称、能源类型、经营单位、经纬度等
- 事件池-主数据充电设备表包含：充电桩编号、充电枪数量、充电类型等
- 事件池-刷卡刷码数据(TA_SJC_SUB_KL_ICQR_MI)包含：刷卡刷码客流数据
- 事件池-主数据车辆表包含：车辆变更事件信息

问题列表：
{q_text}

请输出JSON数组，每个元素包含：row（行号）, H（分类说明）, F（数据表）, G（字段）
格式: [{{"row": 2, "H": "（1）已接入数据可回答", "F": "表名", "G": "字段1,字段2"}}, ...]
只输出JSON，不要其他文字。'''

    payload = {
        "model": "Claude Sonnet 4.6",
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": 8000
    }
    
    cmd = [
        "curl", "-s", "-X", "POST", URL,
        "-H", f"Authorization: Bearer {API_KEY}",
        "-H", "anthropic-version: 2023-06-01",
        "-H", "Content-Type: application/json",
        "-d", json.dumps(payload, ensure_ascii=False)
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        resp = json.loads(result.stdout)
        if "content" in resp and len(resp["content"]) > 0:
            text = resp["content"][0]["text"]
            # Extract JSON array
            start = text.find("[")
            end = text.rfind("]")
            if start >= 0 and end > start:
                json_str = text[start:end+1]
                batch_results = json.loads(json_str)
                for r in batch_results:
                    all_results[r["row"]] = {
                        "H": r.get("H", ""),
                        "F": r.get("F", ""),
                        "G": r.get("G", "")
                    }
                print(f"  Got {len(batch_results)} results", flush=True)
            else:
                print(f"  No JSON array found in response: {text[:200]}", flush=True)
                for q in batch:
                    all_results[q["row"]] = {"H": "（3）数据表schema不支持", "F": "", "G": ""}
        else:
            print(f"  API error: {resp.get('error', resp)}", flush=True)
            for q in batch:
                all_results[q["row"]] = {"H": "（3）数据表schema不支持", "F": "", "G": ""}
    except Exception as e:
        print(f"  Error: {e}", flush=True)
        for q in batch:
            all_results[q["row"]] = {"H": "（3）数据表schema不支持", "F": "", "G": ""}
    
    time.sleep(2)

# Save results
with open("classification_results.json", "w") as f:
    json.dump(all_results, f, ensure_ascii=False, indent=2)

print(f"\n=== Total results: {len(all_results)} ===")
# Print summary
cat_counts = {"（1）": 0, "（2）": 0, "（3）": 0}
for row, r in sorted(all_results.items()):
    h = r["H"]
    if "（1）" in h:
        cat_counts["（1）"] += 1
    elif "（2）" in h:
        cat_counts["（2）"] += 1
    else:
        cat_counts["（3）"] += 1
    print(f"  Row {row}: H={h[:60]} | F={r['F'][:40]} | G={r['G'][:40]}")

print(f"\nCategory counts: (1)={cat_counts['（1）']}, (2)={cat_counts['（2）']}, (3)={cat_counts['（3）']}")
