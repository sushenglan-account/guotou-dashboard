# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('国投测试数据.xlsx')
ws = wb['点击项目数字的弹窗显示列']

# Platform grouping by row ranges (based on merged cells in original template)
# Rows 38-44: 基础设施 (7 rows)
# Rows 45-51: 网络安全 (7 rows)
# Row 52: 专项服务 (1 row)

infra_rows = list(range(38, 45))   # 38-44 inclusive
sec_rows = list(range(45, 52))     # 45-51 inclusive
svc_rows = [52]                    # 52

infra_systems = [
    '存储管理系统',
    '网络建设系统',
    '云平台管理系统',
    '云平台运维系统',
    '网络运维管理系统',
    '数据中心管理系统',
    '存储优化系统',
]

sec_systems = [
    '服务器安全管理系统',
    '数据中心安全管理系统',
    '存储安全管理系统',
    '存储安全防护系统',
    '网络运维安全系统',
    '云平台安全管理系统',
    '系统集成安全管理系统',
]

svc_systems = [
    '综合运维服务系统',
]

# Assign system names
for idx, r in enumerate(infra_rows):
    sys_name = infra_systems[idx]
    ws.cell(row=r, column=2, value=sys_name)
    print(f"  Row {r} (基础设施): B='{sys_name}'")

for idx, r in enumerate(sec_rows):
    sys_name = sec_systems[idx]
    ws.cell(row=r, column=2, value=sys_name)
    print(f"  Row {r} (网络安全): B='{sys_name}'")

for idx, r in enumerate(svc_rows):
    sys_name = svc_systems[idx]
    ws.cell(row=r, column=2, value=sys_name)
    print(f"  Row {r} (专项服务): B='{sys_name}'")

# Also need to update the merged cells for B column if they exist
# B column for these groups might have been merged in original template
# Check and unmerge if needed, then set individual values
for r in infra_rows + sec_rows + svc_rows:
    ws.cell(row=r, column=2, value=ws.cell(row=r, column=2).value)

# Save
output_path = '国投测试数据.xlsx'
wb.save(output_path)
print(f"\nSaved to: {output_path}")

# Verify
wb2 = openpyxl.load_workbook(output_path, data_only=False)
ws2 = wb2['点击项目数字的弹窗显示列']
print("\n=== Verification ===")
for r in range(38, 53):
    a_val = ws2.cell(row=r, column=1).value
    b_val = ws2.cell(row=r, column=2).value
    c_val = ws2.cell(row=r, column=3).value
    e_val = ws2.cell(row=r, column=5).value
    if b_val:
        print(f"  Row {r}: A='{a_val}', B='{b_val}', C='{c_val}', E='{e_val}'")

# Check if any xxx system remains
remaining = 0
for r in range(3, 53):
    b_val = ws2.cell(row=r, column=2).value
    if b_val and 'xxx系统' in str(b_val):
        remaining += 1
        print(f"  STILL HAS xxx_system at row {r}: {b_val}")
if remaining == 0:
    print("\n  All 'xxx系统' replaced successfully!")
else:
    print(f"\n  WARNING: {remaining} 'xxx系统' remain")
